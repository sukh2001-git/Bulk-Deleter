# Copyright (c) 2026, Sukhman and contributors
# For license information, please see license.txt

import frappe
import json
import os
from frappe.model.document import Document
from frappe import _
import openpyxl


class BulkDeleteRequest(Document):
    def validate(self):
        if not self.delete_all_records and not self.excel_file:
            frappe.throw(_("Either upload an Excel file or select 'Delete All Records'"))
        self.total_records = self.get_total_record_count()

    def get_total_record_count(self):
        if self.delete_all_records:
            count = frappe.db.count(self.doctype_name)
            if self.batch_size:
                return min(count, self.batch_size)
            return count
        if self.excel_file and self.file_path:
            try:
                wb = openpyxl.load_workbook(self.file_path, read_only=True)
                ws = wb.active
                count = max(0, ws.max_row - 1)
                if self.batch_size:
                    return min(count, self.batch_size)
                return count
            except:
                return 0
        return 0

    def before_save(self):
        if self.excel_file:
            try:
                file_doc = frappe.get_doc("File", {"file_url": self.excel_file})
                if file_doc:
                    self.file_path = file_doc.get_full_path()
            except:
                pass

    def process_deletions(self):
        frappe.log_error("process_deletions started", f"Bulk Delete Start | {self.name} | doctype: {self.doctype_name} | batch_size: {self.batch_size}")

        self.status = "Processing"
        self.processed_records = 0
        self.successful_deletions = 0
        self.failed_deletions = 0
        self.error_log = ""
        self.save()
        frappe.db.commit()

        try:
            records = self.get_records_to_delete()
            frappe.log_error("Records fetched", f"Bulk Delete | {self.name} | total: {len(records)}")

            if not records:
                frappe.log_error("No records found", f"Bulk Delete | {self.name}")
                self.status = "Completed"
                self.save()
                frappe.db.commit()
                return

            # Handle linkages first if needed
            if self.remove_linkages:
                link_fields_cache = self._get_link_fields()
                frappe.log_error("Link fields fetched", f"Bulk Delete | {self.name} | count: {len(link_fields_cache)}")

                for field in link_fields_cache:
                    try:
                        if self.linkage_strategy == "Delete Links":
                            frappe.db.sql("""
                                DELETE FROM `tab{doctype}`
                                WHERE `{fieldname}` IN %(records)s
                            """.format(
                                doctype=field.get("parent"),
                                fieldname=field.get("fieldname")
                            ), {"records": records})
                        else:
                            frappe.db.sql("""
                                UPDATE `tab{doctype}`
                                SET `{fieldname}` = NULL
                                WHERE `{fieldname}` IN %(records)s
                            """.format(
                                doctype=field.get("parent"),
                                fieldname=field.get("fieldname")
                            ), {"records": records})
                    except Exception as e:
                        frappe.log_error("Linkage handling error", f"Bulk Delete | {self.name} | field: {field} | error: {str(e)}")

                frappe.db.commit()
                frappe.log_error("Linkages handled", f"Bulk Delete | {self.name}")

            # Delete all records in one SQL query
            frappe.db.sql("""
                DELETE FROM `tab{doctype}`
                WHERE name IN %(records)s
            """.format(doctype=self.doctype_name), {"records": records})

            frappe.db.commit()
            frappe.log_error("Bulk delete done", f"Bulk Delete | {self.name} | deleted: {len(records)}")

            self.processed_records = len(records)
            self.successful_deletions = len(records)
            self.status = "Completed"

            # Single bulk log insert
            log_buffer = [
                self._make_log_entry(r, "Success", "") for r in records
            ]
            self._flush_logs(log_buffer)

            self.save()
            frappe.db.commit()

        except Exception as e:
            frappe.log_error("process_deletions exception", frappe.get_traceback())
            self.status = "Failed"
            self.error_log = str(e)
            self.save()
            frappe.db.commit()
            raise

    def get_records_to_delete(self):
        frappe.log_error("get_records_to_delete started", f"Bulk Delete | {self.name} | delete_all: {self.delete_all_records} | batch_size: {self.batch_size}")

        if self.delete_all_records:
            query = """
                SELECT name FROM `tab{doctype}`
                WHERE name != 'Administrator'
            """.format(doctype=self.doctype_name)

            if self.batch_size:
                query += " LIMIT {0}".format(int(self.batch_size))

            frappe.log_error("Executing delete all query", f"Bulk Delete | {self.name} | query: {query}")
            records = frappe.db.sql_list(query)
            frappe.log_error("Delete all records fetched", f"Bulk Delete | {self.name} | count: {len(records)}")
            return records

        records = []

        if not self.file_path or not os.path.exists(self.file_path):
            frappe.log_error("File path missing or not found", f"Bulk Delete | {self.name} | file_path: {self.file_path} | excel_file: {self.excel_file}")
            if self.excel_file:
                try:
                    file_doc = frappe.get_doc("File", {"file_url": self.excel_file})
                    self.file_path = file_doc.get_full_path()
                    frappe.log_error("File path resolved", f"Bulk Delete | {self.name} | resolved_path: {self.file_path}")
                except:
                    frappe.log_error("File path resolution failed", f"Bulk Delete | {self.name} | excel_file: {self.excel_file}")
                    frappe.throw(_("Could not find uploaded file"))

        try:
            frappe.log_error("Opening Excel file", f"Bulk Delete | {self.name} | file_path: {self.file_path}")
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            ws = wb.active
            frappe.log_error("Excel opened", f"Bulk Delete | {self.name} | max_row: {ws.max_row} | max_col: {ws.max_column}")

            header_row = list(ws)[0]
            headers = [str(cell.value).strip() if cell.value else "" for cell in header_row]
            frappe.log_error("Headers found", f"Bulk Delete | {self.name} | headers: {headers} | looking_for: {self.name_column}")

            col_idx = None
            for idx, cell in enumerate(header_row):
                if cell.value and str(cell.value).strip().lower() == str(self.name_column).strip().lower():
                    col_idx = idx
                    break

            if col_idx is None:
                frappe.log_error("Column not found", f"Bulk Delete | {self.name} | name_column: {self.name_column} | available_headers: {headers}")
                frappe.throw(_("Column '{0}' not found in Excel").format(self.name_column))

            frappe.log_error("Column found", f"Bulk Delete | {self.name} | col_idx: {col_idx} | name_column: {self.name_column}")

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx + 1)
                if cell.value:
                    records.append(str(cell.value).strip())

            frappe.log_error("Records read from Excel", f"Bulk Delete | {self.name} | total_read: {len(records)}")

            if self.batch_size:
                records = records[:int(self.batch_size)]
                frappe.log_error("Batch size applied", f"Bulk Delete | {self.name} | after_limit: {len(records)}")

        except frappe.exceptions.ValidationError:
            raise
        except Exception as e:
            frappe.log_error("Excel read error", f"Bulk Delete | {self.name} | error: {frappe.get_traceback()}")
            frappe.throw(_("Error reading Excel: {0}").format(str(e)))

        frappe.log_error("get_records_to_delete completed", f"Bulk Delete | {self.name} | final_count: {len(records)}")
        return records

    def _get_link_fields(self):
        try:
            return frappe.get_all(
                "DocField",
                filters={
                    "options": self.doctype_name,
                    "fieldtype": ["in", ("Link", "Dynamic Link")]
                },
                fields=["parent", "fieldname"]
            )
        except:
            return []

    def _find_linked_documents(self, record_name, link_fields_cache):
        linked = []
        for field in link_fields_cache:
            try:
                results = frappe.db.sql("""
                    SELECT name FROM `tab{doctype}`
                    WHERE `{fieldname}` = %s
                """.format(
                    doctype=field.get("parent"),
                    fieldname=field.get("fieldname")
                ), record_name, as_dict=True)

                for row in results:
                    linked.append({
                        "doctype": field.get("parent"),
                        "name": row.name,
                        "field": field.get("fieldname")
                    })
            except Exception as e:
                frappe.log_error("Find linked docs error", f"Bulk Delete | field: {field} | error: {str(e)}")
                continue
        return linked

    def _make_log_entry(self, record_name, status, reason, linked_docs=None):
        return {
            "name": frappe.generate_hash(length=10),
            "request": self.name,
            "doctype_name": self.doctype_name,
            "record_name": record_name,
            "status": status,
            "reason": reason,
            "linked_docs": json.dumps(linked_docs) if linked_docs else "",
            "deleted_by": frappe.session.user,
            "creation": frappe.utils.now(),
            "modified": frappe.utils.now(),
            "owner": frappe.session.user,
            "modified_by": frappe.session.user,
        }

    def _flush_logs(self, log_buffer):
        if not log_buffer:
            return
        try:
            frappe.db.bulk_insert(
                "Bulk Delete Log",
                fields=list(log_buffer[0].keys()),
                values=[list(entry.values()) for entry in log_buffer],
                ignore_duplicates=True
            )
        except Exception as e:
            frappe.log_error("Log flush error", f"Bulk Delete | error: {str(e)}")


def run_deletion_job(docname):
    """Background job to process deletions"""
    doc = frappe.get_doc("Bulk Delete Request", docname)
    doc.process_deletions()